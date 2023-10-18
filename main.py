import os
import csv
import openpyxl
#from docx import Document
#import docx2txt
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext


# Function to search for a query in a CSV file
def search_csv(file_path, query):
    results = []
    with open(file_path, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            for item in row:
                if query.lower() in item.lower():
                    results.append(row)
                    break
    return results


# Function to search for a query in an XLSX file
def search_xlsx(file_path, query):
    results = []
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if query.lower() in str(cell).lower():
                    results.append(row)
                    break
    return results


# Function to search for a query in a TXT file
def search_txt(file_path, query):
    results = []
    with open(file_path, 'r') as file:
        for line in file.readlines():
            if query.lower() in line.lower():
                results.append(line.strip())
    return results




# Function to handle the search button click event
def search_button_clicked():
    query = query_entry.get()
    directory = directory_entry.get()
    file_types = file_types_entry.get().split(",")
    search_results_text.delete(1.0, tk.END)

    if query == "":
        messagebox.showwarning("Empty Query", "Please enter a search query.")
        return

    if directory == "":
        messagebox.showwarning("Empty Directory", "Please select a directory to search in.")
        return

    if len(file_types) == 0:
        messagebox.showwarning("No File Types", "Please enter at least one file type to search in.")
        return

    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)

            if file.lower().endswith(tuple(file_types)):
                if file.lower().endswith('.csv'):
                    search_results = search_csv(file_path, query)
                elif file.lower().endswith('.xlsx'):
                    search_results = search_xlsx(file_path, query)
                elif file.lower().endswith('.txt'):
                    search_results = search_txt(file_path, query)
                #elif file.lower().endswith('.docx'):
                    #search_results = search_docx(file_path, query)
                #elif file.lower().endswith('.doc'):
                    #search_results = search_doc(file_path, query)
                else:
                    continue

                if len(search_results) > 0:
                    search_results_text.insert(tk.END, "File: " + file_path + "\n")
                    search_results_text.insert(tk.END, "Search results:\n")
                    for result in search_results:
                        search_results_text.insert(tk.END, "- " + str(result) + "\n")
                    search_results_text.insert(tk.END, "\n")


# Create the main Tkinter window
window = tk.Tk()
window.title("File Search App")

# Create and position the widgets
query_label = tk.Label(window, text="Search Query:")
query_label.grid(row=0, column=0, padx=10, pady=10)
query_entry = tk.Entry(window)
query_entry.grid(row=0, column=1, padx=10, pady=10)

directory_label = tk.Label(window, text="Directory:")
directory_label.grid(row=1, column=0, padx=10, pady=10)
directory_entry = tk.Entry(window)
directory_entry.grid(row=1, column=1, padx=10, pady=10)
directory_button = tk.Button(window, text="Browse",
                             command=lambda: directory_entry.insert(tk.END, filedialog.askdirectory()))
directory_button.grid(row=1, column=2, padx=10, pady=10)

file_types_label = tk.Label(window, text="File Types (comma-separated):")
file_types_label.grid(row=2, column=0, padx=10, pady=10)
file_types_entry = tk.Entry(window)
file_types_entry.grid(row=2, column=1, padx=10, pady=10)

search_button = tk.Button(window, text="Search", command=search_button_clicked)
search_button.grid(row=3, column=0, columnspan=3, padx=10, pady=10)

search_results_text = scrolledtext.ScrolledText(window, width=80, height=20)
search_results_text.grid(row=4, column=0, columnspan=3, padx=10, pady=10)

# Run the main Tkinter event loop
window.mainloop()